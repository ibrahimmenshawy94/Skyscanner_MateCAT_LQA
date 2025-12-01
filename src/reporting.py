import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def format_tb_matches_friendly(matches_list):
    """
    Converts a list of dicts [{'source': 'A', 'target': 'B'}]
    into a string "A -> B\nC -> D".
    """
    if not isinstance(matches_list, list) or not matches_list:
        return "No matches"

    lines = []
    for m in matches_list:
        src = m.get("source") or m.get("src")
        trg = m.get("target") or m.get("trg")
        if src and trg:
            lines.append(f"• {src} → {trg}")

    return "\n".join(lines) if lines else "No matches"


def generate_lqa_scorecard(df_result, base_output_folder, language_name):
    """
    Generates a formatted Excel scorecard from the LQA results.
    Adds HITL columns (green headers) with dropdowns for Status and Category-Subcategory.
    """
    lang_folder = os.path.join(base_output_folder, language_name)
    os.makedirs(lang_folder, exist_ok=True)

    file_name = f"MosAIQ LQA_{language_name}.xlsx"
    output_path = os.path.join(lang_folder, file_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "LQA Scorecard"

    static_cols_config = [
        ("Job ID", "Job_ID", 12, False),
        ("Segment ID", "Segment_ID", 12, False),
        ("Char Limit", "Character_Limit", 12, False),
        ("Link", "Link", 12, True),
        ("Note", "Note", 25, True),
        ("Context Before", "Context_Before", 25, True),
        ("Context After", "Context_After", 25, True),
        ("Source", "Source", 40, True),
        ("Target", "Target", 40, True),
        ("TB Matches", "TB Matches", 25, True),
    ]

    error_headers = ["AI Category - Sub Category", "AI Severity", "AI Rationale"]
    error_widths = [20, 20, 30]

    hitl_headers = ["Status", "HITL Category - Sub Category", "HITL Severity", "HITL Rationale", "HITL Final Target"]
    hitl_widths = [12, 25, 15, 30, 40]

    final_cols_config = [
        ("Final Target", "Final_Target", 40, True),
        ("Char Count", None, 15, False),
    ]

    all_headers = [c[0] for c in static_cols_config] + error_headers + [c[0] for c in final_cols_config] + hitl_headers
    ws.append(all_headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    hitl_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    thin_border = Side(border_style="thin", color="000000")
    thick_border = Side(border_style="medium", color="000000")

    for col_num, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_num)
        if header in hitl_headers:
            cell.fill = hitl_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = ws.dimensions

    current_row = 2

    for _, row in df_result.iterrows():
        errors = row.get("Final_Errors", [])
        if not isinstance(errors, list):
            errors = []
        has_errors = len(errors) > 0

        char_limit_val = row.get("Character_Limit")
        has_char_limit = False
        if pd.notna(char_limit_val) and str(char_limit_val).strip() != "":
            has_char_limit = True

        num_sub_rows = max(1, len(errors))
        start_row = current_row
        end_row = current_row + num_sub_rows - 1

        col_idx = 1
        for header, df_col, width, wrap in static_cols_config:
            val = row.get(df_col, "")

            if df_col == "TB Matches":
                val = format_tb_matches_friendly(val)

            cell = ws.cell(row=start_row, column=col_idx, value=val)

            if df_col == "Link" and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = link_font

            if num_sub_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            col_idx += 1

        # AI Errors
        error_col_start = col_idx

        if has_errors:
            for i in range(num_sub_rows):
                r = start_row + i
                cat_val = sev_val = rat_val = ""

                if i < len(errors):
                    err = errors[i]
                    main_cat = err.get("category", "")
                    sub_cat = err.get("subcategory", "")
                    cat_val = f"{main_cat} - {sub_cat}" if sub_cat else main_cat
                    sev_val = err.get("severity", "")
                    rat_val = err.get("rationale", "")

                ws.cell(row=r, column=error_col_start, value=cat_val).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
                ws.cell(row=r, column=error_col_start + 1, value=sev_val).alignment = Alignment(
                    vertical="top", wrap_text=True
                )
                ws.cell(row=r, column=error_col_start + 2, value=rat_val).alignment = Alignment(
                    vertical="top", wrap_text=True
                )

        col_idx += 3

        # Final Target & Count (only when errors exist; otherwise blank but merged for alignment)
        if has_errors:
            final_target_val = row.get("Final_Target", "")
            ft_cell = ws.cell(row=start_row, column=col_idx, value=final_target_val)
            if num_sub_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            ft_cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            ft_cell = ws.cell(row=start_row, column=col_idx, value=None)
            if num_sub_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            ft_cell.alignment = Alignment(vertical="top", wrap_text=True)

        col_idx += 1

        if has_errors and has_char_limit:
            target_cell_ref = ft_cell.coordinate
            formula = f"=LEN({target_cell_ref})"
            count_cell = ws.cell(row=start_row, column=col_idx, value=formula)
            if num_sub_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            count_cell.alignment = Alignment(vertical="top", horizontal="center")
        else:
            count_cell = ws.cell(row=start_row, column=col_idx, value=None)
            if num_sub_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            count_cell.alignment = Alignment(vertical="top", horizontal="center")

        col_idx += 1

        # HITL Columns (no merging; per-error rows editable)
        hitl_status_col = col_idx
        hitl_cat_col = col_idx + 1
        hitl_sev_col = col_idx + 2
        hitl_rat_col = col_idx + 3
        hitl_final_col = col_idx + 4

        for i in range(num_sub_rows):
            r = start_row + i
            ws.cell(row=r, column=hitl_status_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=hitl_cat_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=hitl_sev_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=hitl_rat_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=hitl_final_col).alignment = Alignment(vertical="top", wrap_text=True)

        col_idx += 5

        total_cols = len(static_cols_config) + len(error_headers) + len(final_cols_config) + len(hitl_headers)

        for r in range(start_row, end_row + 1):
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=r, column=c)

                top = thin_border if r > start_row else thick_border
                bottom = thin_border if r < end_row else thick_border
                left = thin_border
                right = thin_border

                if c == 1:
                    left = thick_border
                if c == total_cols:
                    right = thick_border

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

        current_row += num_sub_rows

    # Column widths
    current_col = 1
    for _, _, width, _ in static_cols_config:
        ws.column_dimensions[get_column_letter(current_col)].width = width
        current_col += 1
    for width in error_widths:
        ws.column_dimensions[get_column_letter(current_col)].width = width
        current_col += 1
    for _, _, width, _ in final_cols_config:
        ws.column_dimensions[get_column_letter(current_col)].width = width
        current_col += 1
    for width in hitl_widths:
        ws.column_dimensions[get_column_letter(current_col)].width = width
        current_col += 1

    limit_col_letter = "C"
    # Char Count column index: static + error + 2nd of final_cols_config
    count_col_letter = get_column_letter(len(static_cols_config) + len(error_headers) + 2)
    format_range = f"{count_col_letter}2:{count_col_letter}{max(2, current_row-1)}"

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")

    red_formula = f'AND(ISNUMBER({count_col_letter}2), ${limit_col_letter}2<>"", {count_col_letter}2>${limit_col_letter}2)'
    ws.conditional_formatting.add(format_range, FormulaRule(formula=[red_formula], stopIfTrue=True, fill=red_fill, font=red_font))

    green_formula = f'AND(ISNUMBER({count_col_letter}2), ${limit_col_letter}2<>"", {count_col_letter}2<=${limit_col_letter}2)'
    ws.conditional_formatting.add(format_range, FormulaRule(formula=[green_formula], stopIfTrue=True, fill=green_fill, font=green_font))

    # Data validations for HITL columns
    if current_row > 2:
        last_row = current_row - 1
        status_col_letter = get_column_letter(len(static_cols_config) + len(error_headers) + len(final_cols_config) + 1)
        cat_col_letter = get_column_letter(len(static_cols_config) + len(error_headers) + len(final_cols_config) + 2)

        status_dv = DataValidation(type="list", formula1='"Accepted,Rejected,Edited"', allow_blank=True)
        status_dv.add(f"{status_col_letter}2:{status_col_letter}{last_row}")
        ws.add_data_validation(status_dv)

        hitl_sev_options = ["Critical", "Major", "Minor", "Neutral"]
        sev_list = ",".join(hitl_sev_options)
        hitl_sev_col_letter = get_column_letter(len(static_cols_config) + len(error_headers) + len(final_cols_config) + 3)
        sev_dv = DataValidation(type="list", formula1=f'"{sev_list}"', allow_blank=True)
        sev_dv.add(f"{hitl_sev_col_letter}2:{hitl_sev_col_letter}{last_row}")
        ws.add_data_validation(sev_dv)

        hitl_cat_options = [
            "Accuracy - Mistranslation",
            "Accuracy - Omission",
            "Accuracy - Addition",
            "Terminology - Term inconsistency",
            "Grammar - Syntax error",
            "Grammar - Morphology error",
            "Style - Tone/register",
            "Style - Redundancy",
            "Style - Ambiguity",
            "Locale Conventions - Punctuation",
            "Locale Conventions - Date format",
            "Locale Conventions - Time format",
            "Locale Conventions - Number format",
            "Formatting - Placeholder mismatch",
            "Formatting - Tag mismatch",
        ]
        cat_list = ",".join(hitl_cat_options)
        cat_dv = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
        cat_dv.add(f"{cat_col_letter}2:{cat_col_letter}{last_row}")
        ws.add_data_validation(cat_dv)

        # Expand validation range for HITL severity and status across all error sub-rows
        status_dv.ranges = []
        status_dv.add(f"{status_col_letter}2:{status_col_letter}{last_row}")
        sev_dv.ranges = []
        sev_dv.add(f"{hitl_sev_col_letter}2:{hitl_sev_col_letter}{last_row}")

    wb.save(output_path)
    print(f"Scorecard saved to: {output_path}")
    return output_path
